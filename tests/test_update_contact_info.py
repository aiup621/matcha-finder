from bs4 import BeautifulSoup
from pathlib import Path
import sys
import requests

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import update_contact_info as uc
from update_contact_info_api import select_best_email


def test_find_instagram():
    html = '<a href="https://www.instagram.com/test">IG</a>'
    soup = BeautifulSoup(html, "html.parser")
    assert uc.find_instagram(soup, "http://example.com") == "https://www.instagram.com/test"


def test_crawl_site_for_email_from_mailto(monkeypatch):
    pages = {"http://example.com": '<a href="mailto:info@example.com">mail</a>'}

    def fake_fetch(url, timeout=5, verify=True):
        return pages.get(url)

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert uc.crawl_site_for_email("http://example.com") == "info@example.com"


def test_crawl_site_for_email_normalizes(monkeypatch):
    pages = {
        "http://example.com": "<a href='/next'>next</a>",
        "http://example.com/next": "Contact: sales[at]example.com",
    }

    def fake_fetch(url, timeout=5, verify=True):
        return pages.get(url)

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert (
        uc.crawl_site_for_email("http://example.com", max_depth=2)
        == "sales@example.com"
    )


def test_crawl_site_for_email_unescapes(monkeypatch):
    pages = {"http://example.com": "Contact: info&#64;example.com"}

    def fake_fetch(url, timeout=5, verify=True):
        return pages.get(url)

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert uc.crawl_site_for_email("http://example.com") == "info@example.com"


def test_crawl_site_for_email_skips_blocklisted(monkeypatch):
    pages = {
        "http://example.com": (
            "<a href='mailto:career@example.com'>c</a>"
            "<a href='mailto:info@example.com'>i</a>"
        )
    }

    def fake_fetch(url, timeout=5, verify=True):
        return pages.get(url)

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert uc.crawl_site_for_email("http://example.com") == "info@example.com"


def test_crawl_site_for_email_returns_none_if_only_blocklisted(monkeypatch):
    pages = {"http://example.com": "<a href='mailto:catering@example.com'>c</a>"}

    def fake_fetch(url, timeout=5, verify=True):
        return pages.get(url)

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert uc.crawl_site_for_email("http://example.com") is None


def test_find_contact_form(monkeypatch):
    html = '<a href="/contact">contact</a>'
    soup = BeautifulSoup(html, "html.parser")

    def fake_fetch(url, timeout=5, verify=True):
        return "<form></form>" if url.endswith("/contact") else None

    monkeypatch.setattr(uc, "_fetch_page", fake_fetch)
    assert (
        uc.find_contact_form(soup, "http://example.com")
        == "http://example.com/contact"
    )


def test_fetch_page_retries_on_403(monkeypatch):
    calls = []

    class Resp:
        def __init__(self, status):
            self.status_code = status
            self.text = "ok"

        def raise_for_status(self):
            if self.status_code >= 400:
                raise requests.HTTPError(response=self)

    def fake_get(url, timeout, verify=True, headers=None):
        calls.append(headers.get("User-Agent"))
        return Resp(403) if len(calls) == 1 else Resp(200)

    monkeypatch.setattr(uc.requests, "get", fake_get)
    assert uc._fetch_page("http://example.com", timeout=5) == "ok"
    assert len(calls) == 2
    assert "Mozilla" in calls[0]


def test_process_sheet_row_range(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://a")
    ws.cell(row=3, column=3, value="http://b")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=2, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"
    assert ws2.cell(row=3, column=7).value is None


def test_process_specific_worksheet(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.cell(row=2, column=1, value="ok")
    ws1.cell(row=2, column=3, value="http://a")
    ws2 = wb.create_sheet("Sheet2")
    ws2.cell(row=2, column=1, value="ok")
    ws2.cell(row=2, column=3, value="http://b")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=2, worksheet="Sheet2")

    wb2 = openpyxl.load_workbook(file)
    assert wb2["Sheet1"].cell(row=2, column=7).value is None
    assert wb2["Sheet2"].cell(row=2, column=7).value == "なし"


def test_stop_on_blank_column_a(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://a")
    ws.cell(row=3, column=3, value="http://b")  # A3 is blank
    ws.cell(row=4, column=1, value="ok")
    ws.cell(row=4, column=3, value="http://c")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"
    assert ws2.cell(row=3, column=7).value is None
    assert ws2.cell(row=4, column=7).value is None


def test_start_row_ignores_action_end(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=1, column=1, value="Action")
    ws.cell(row=1, column=3, value=2)  # would normally stop at row 2
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://a")
    ws.cell(row=3, column=1, value="ok")
    ws.cell(row=3, column=3, value="http://b")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"
    assert ws2.cell(row=3, column=7).value == "なし"


def test_skip_invalid_url(tmp_path, monkeypatch):
    import openpyxl

    calls = []

    def dummy_get(url, timeout, verify=True, headers=None):
        calls.append(url)
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="not a url")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=2, worksheet="Sheet")

    assert calls == []
    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value is None


def test_process_sheet_from_url(tmp_path, monkeypatch):
    import io
    import openpyxl

    # Create an in-memory workbook to be served over a faux HTTP request.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://page")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    class WorkbookResponse:
        def __init__(self, data):
            self.content = data

        def raise_for_status(self):
            pass

    class PageResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass
        status_code = 200

        def raise_for_status(self):
            pass

    def fake_get(url, timeout=10, verify=True, headers=None):
        return WorkbookResponse(content) if url == "http://sheet" else PageResponse()

    monkeypatch.setattr(uc.requests, "get", fake_get)
    monkeypatch.chdir(tmp_path)
    uc.process_sheet("http://sheet", start_row=2, end_row=2, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(tmp_path / "downloaded.xlsx")
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"


def test_google_sheet_link_is_transformed(monkeypatch, tmp_path):
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://page")
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    class WorkbookResponse:
        def __init__(self, content):
            self.content = content

        def raise_for_status(self):
            pass

    class PageResponse:
        text = "<html></html>"
        status_code = 200

        def raise_for_status(self):
            pass

    def fake_get(url, timeout=10, verify=True, headers=None):
        if url.startswith(
            "https://docs.google.com/spreadsheets/d/FILEID/export?format=xlsx&gid=0"
        ):
            return WorkbookResponse(data)
        if url == "http://page":
            return PageResponse()
        raise AssertionError(f"unexpected url {url}")

    monkeypatch.setattr(uc.requests, "get", fake_get)
    monkeypatch.chdir(tmp_path)
    uc.process_sheet(
        "https://docs.google.com/spreadsheets/d/FILEID/edit?gid=0#gid=0",
        start_row=2,
        end_row=2,
        worksheet="Sheet",
    )

    wb2 = openpyxl.load_workbook(tmp_path / "downloaded.xlsx")
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"


def test_retry_on_ssl_error(tmp_path, monkeypatch):
    import openpyxl
    from requests.exceptions import SSLError

    calls = []

    def dummy_get(url, timeout, verify=True, headers=None):
        calls.append(verify)
        if verify:
            raise SSLError("bad cert")

        class DummyResponse:
            text = "<html></html>"
            status_code = 200

            def raise_for_status(self):
                pass

        return DummyResponse()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="https://a")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=2, worksheet="Sheet")

    assert calls[:2] == [True, False]
    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=2, column=7).value == "なし"


def test_request_failure_marks_error(tmp_path, monkeypatch):
    import openpyxl
    from requests.exceptions import ConnectionError

    def dummy_get(url, timeout, verify=True, headers=None):
        raise ConnectionError("fail")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="https://a")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=2, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    assert ws2.max_row == 1


def test_duplicate_emails_are_removed(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        def __init__(self, text):
            self.text = text
            self.status_code = 200

        def raise_for_status(self):
            pass

    pages = {
        "http://a": "<a href='mailto:info@example.com'>mail</a>",
        "http://b": "<a href='mailto:INFO@example.com'>mail</a>",
    }

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse(pages.get(url, "<html></html>"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://a")
    ws.cell(row=3, column=1, value="ok")
    ws.cell(row=3, column=3, value="http://b")
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=3, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    emails = [
        ws2.cell(row=row, column=5).value
        for row in range(2, ws2.max_row + 1)
        if ws2.cell(row=row, column=5).value
    ]
    assert emails == ["info@example.com"]


def test_duplicate_hyperlink_emails_are_removed(tmp_path, monkeypatch):
    import openpyxl

    class DummyResponse:
        def __init__(self, text):
            self.text = text
            self.status_code = 200

        def raise_for_status(self):
            pass

    def dummy_get(url, timeout, verify=True, headers=None):
        return DummyResponse("<html></html>")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.cell(row=2, column=1, value="ok")
    ws.cell(row=2, column=3, value="http://a")
    ws.cell(row=2, column=5, value="info@example.com")
    ws.cell(
        row=3,
        column=1,
        value="ok",
    )
    ws.cell(row=3, column=3, value="http://b")
    ws.cell(
        row=3,
        column=5,
        value='=HYPERLINK("mailto:INFO@example.com","Contact")',
    )
    file = tmp_path / "sample.xlsx"
    wb.save(file)

    monkeypatch.setattr(uc.requests, "get", dummy_get)
    uc.process_sheet(str(file), start_row=2, end_row=3, worksheet="Sheet")

    wb2 = openpyxl.load_workbook(file)
    ws2 = wb2["Sheet"]
    emails = [
        ws2.cell(row=row, column=5).value
        for row in range(2, ws2.max_row + 1)
        if ws2.cell(row=row, column=5).value
    ]
    assert emails == ["info@example.com"]


def test_normalize_email_handles_mailto_formula():
    assert (
        uc._normalize_email('=HYPERLINK("mailto:Info@Example.com","Contact")')
        == "info@example.com"
    )


def test_normalize_email_prefers_hyperlink_target():
    assert (
        uc._normalize_email("お問い合わせはこちら", "mailto:Sales@Example.com")
        == "sales@example.com"
    )


def test_orders_blocked_on_consumer_context():
    c = [{'email':'orders@cafe.com','source_url':'https://cafe.com/menu','anchor_text':'Order online'}]
    best, notes, kept, blocked = select_best_email(c, 'https://cafe.com')
    assert best is None
    assert any('consumer_order_inbox' in b[1] for b in blocked)


def test_orders_allowed_on_trade_context():
    c = [{'email':'orders@cafe.com','source_url':'https://cafe.com/wholesale','anchor_text':'Wholesale orders'}]
    best, notes, kept, blocked = select_best_email(c, 'https://cafe.com')
    assert best == 'orders@cafe.com'


def test_catering_address_blocked():
    c = [{'email':'info@mycatering.com','source_url':'https://cafe.com'}]
    best, notes, kept, blocked = select_best_email(c, 'https://cafe.com')
    assert best is None
    assert any('catering' in b[1] for b in blocked)


def test_process_sheet_download_error(monkeypatch):
    import requests

    def bad_get(url):
        raise requests.exceptions.ConnectionError('fail')

    monkeypatch.setattr(uc.requests, 'get', bad_get)
    uc.process_sheet('http://sheet', start_row=2, worksheet='Sheet')
