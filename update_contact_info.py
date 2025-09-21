import argparse
import html
import logging
import re
from collections import deque
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
REQUEST_TIMEOUT = 5
EMAIL_BLOCKLIST = ("catering", "career")

DEFAULT_SHEET_PATH = (
    "https://docs.google.com/spreadsheets/d/1HU-GqN7sBcORIZrYEw4FkyfNmgDtXsO7CtDLVHEsldA/"
    "edit?gid=159511499#gid=159511499"
)


def _fetch_page(url, timeout=REQUEST_TIMEOUT, verify=True):
    """Return the page text with a browser-like ``User-Agent``.

    The function retries on HTTP 403 responses and SSL errors.  When an SSL
    error occurs the certificate verification is disabled for the retry."""

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    for _ in range(3):
        try:
            res = requests.get(url, timeout=timeout, verify=verify, headers=headers)
            if res.status_code == 403:
                continue
            res.raise_for_status()
            return res.text
        except requests.exceptions.SSLError:
            if verify:
                verify = False
                continue
        except requests.RequestException:
            continue
    return None


def find_instagram(soup, base_url):
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "instagram.com" in href:
            return href if href.startswith("http") else urljoin(base_url, href)
    return None


def crawl_site_for_email(base_url, max_depth=1, timeout=REQUEST_TIMEOUT, verify=True):
    """Crawl ``base_url`` breadth-first looking for an email address."""

    parsed = urlparse(base_url)
    domain = parsed.netloc
    queue = deque([(base_url, 0)])
    visited = set()

    while queue:
        url, depth = queue.popleft()
        if url in visited or depth > max_depth:
            continue
        visited.add(url)

        content = _fetch_page(url, timeout=timeout, verify=verify)
        if not content:
            continue

        soup = BeautifulSoup(content, "html.parser")

        mailtos = soup.find_all("a", href=lambda h: h and h.lower().startswith("mailto:"))
        for m in mailtos:
            href = m["href"]
            candidate = re.sub(r"^mailto:", "", href, flags=re.I).split("?")[0]
            if any(b in candidate.lower() for b in EMAIL_BLOCKLIST):
                continue
            return candidate

        text = html.unescape(soup.get_text(" "))
        for pattern in ["[at]", "(at)", "＠"]:
            text = text.replace(pattern, "@")
        for match in EMAIL_RE.finditer(text):
            candidate = match.group(0)
            if any(b in candidate.lower() for b in EMAIL_BLOCKLIST):
                continue
            return candidate

        if depth < max_depth:
            for a in soup.find_all("a", href=True):
                link = urljoin(url, a["href"])
                if urlparse(link).netloc == domain and link not in visited:
                    queue.append((link, depth + 1))
    return None


def find_contact_form(soup, base_url, timeout=REQUEST_TIMEOUT, verify=True):
    candidates = []
    keywords = ["contact", "お問い合わせ", "お問合せ", "inquiry"]
    for a in soup.find_all("a", href=True):
        text = (a.get_text() or "").lower()
        href = a["href"]
        if any(k in text for k in keywords) or any(
            k in href.lower() for k in keywords + ["form"]
        ):
            full = href if href.startswith("http") else urljoin(base_url, href)
            candidates.append(full)

    for link in candidates:
        content = _fetch_page(link, timeout=timeout, verify=verify)
        if not content:
            continue
        if BeautifulSoup(content, "html.parser").find("form"):
            return link
    return None


def _normalize_email(value):
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
    else:
        normalized = str(value).strip()
    if not normalized:
        return None
    return normalized.lower()


def _collect_rows_to_delete(ws, first_data_row=2):
    rows_to_delete = set()
    seen_emails = {}

    for row in range(first_data_row, ws.max_row + 1):
        email_value = _normalize_email(ws.cell(row=row, column=5).value)
        if email_value:
            if email_value in seen_emails:
                rows_to_delete.add(row)
            else:
                seen_emails[email_value] = row

        status = ws.cell(row=row, column=7).value
        if isinstance(status, str) and status.strip() == "エラー":
            rows_to_delete.add(row)

    return rows_to_delete


def process_sheet(path, start_row=None, end_row=None, worksheet="抹茶営業リスト（カフェ）", debug=False):
    import io
    import urllib.parse
    import openpyxl

    if debug:
        logging.basicConfig(level=logging.INFO)

    # URLが来た場合はダウンロードして BytesIO から読み込む。
    # Google Sheets の場合は export エンドポイントに書き換え。
    save_path = path
    if isinstance(path, str) and path.startswith("http"):
        download_url = path
        if "docs.google.com/spreadsheets" in path and "/export" not in path:
            parsed = urllib.parse.urlparse(path)
            match = re.search(r"/d/([^/]+)", parsed.path)
            if match:
                file_id = match.group(1)
                qs = urllib.parse.parse_qs(parsed.query)
                gid = qs.get("gid", ["0"])[0]
                download_url = (
                    f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx&gid={gid}"
                )
        try:
            resp = requests.get(download_url)
            resp.raise_for_status()
        except requests.RequestException as exc:
            logging.error("Failed to download spreadsheet %s: %s", download_url, exc)
            return
        path = io.BytesIO(resp.content)
        save_path = "downloaded.xlsx"

    wb = openpyxl.load_workbook(path)
    ws = wb[worksheet]

    # Determine start and end rows.  If both are unspecified and the sheet
    # contains ``Action`` metadata in the first row, honour those values.
    # Otherwise, default to starting from row 2 and processing until the last
    # row in the sheet.  This ensures that explicitly providing ``start_row``
    # via the command line causes the script to keep scanning downward until
    # column A becomes blank, regardless of any value in ``C1``.
    if start_row is None and end_row is None and ws["A1"].value == "Action":
        try:
            start_row = int(ws["B1"].value)
        except (TypeError, ValueError):
            start_row = 2
        try:
            end_row = int(ws["C1"].value)
        except (TypeError, ValueError):
            end_row = ws.max_row
    else:
        if start_row is None:
            start_row = 2
        if end_row is None:
            end_row = ws.max_row

    end_row = min(end_row, ws.max_row)

    for row in range(start_row, end_row + 1):
        # A列が空なら以降は処理しない
        if not ws.cell(row=row, column=1).value:
            break
        url = ws.cell(row=row, column=3).value
        if not isinstance(url, str):
            continue
        url = url.strip()
        if not url.lower().startswith(("http://", "https://")):
            logging.warning("Skipping invalid URL at row %s: %r", row, url)
            continue
        logging.info("Processing row %s: %s", row, url)
        content = _fetch_page(url, timeout=REQUEST_TIMEOUT)
        if content is None:
            ws.cell(row=row, column=7).value = "エラー"
            continue
        soup = BeautifulSoup(content, "html.parser")
        insta = find_instagram(soup, url)
        email = crawl_site_for_email(url, timeout=REQUEST_TIMEOUT)
        form = find_contact_form(soup, url, timeout=REQUEST_TIMEOUT)
        if insta:
            ws.cell(row=row, column=4).value = insta
        if email:
            ws.cell(row=row, column=5).value = email
        if form:
            ws.cell(row=row, column=6).value = form
        if not any([insta, email, form]):
            ws.cell(row=row, column=7).value = "なし"
        logging.info(
            "Row %s result - Insta: %s, Email: %s, Form: %s",
            row, bool(insta), bool(email), bool(form)
        )

    rows_to_delete = _collect_rows_to_delete(ws)
    if rows_to_delete:
        logging.info("Deleting %d rows due to duplicate emails or errors", len(rows_to_delete))
        for row in sorted(rows_to_delete, reverse=True):
            logging.debug("Deleting row %s", row)
            ws.delete_rows(row)
    wb.save(save_path)


def main():
    parser = argparse.ArgumentParser(description="Update contact info from homepage URLs.")
    parser.add_argument(
        "sheet",
        nargs="?",
        default=DEFAULT_SHEET_PATH,
        help="Path to Excel file to update",
    )
    parser.add_argument("--start-row", type=int, default=None, help="Row number to start processing")
    parser.add_argument("--end-row", type=int, default=None, help="Row number to stop processing (inclusive)")
    parser.add_argument("--worksheet", default="抹茶営業リスト（カフェ）", help="Worksheet name to process")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()
    process_sheet(args.sheet, args.start_row, args.end_row, args.worksheet, args.debug)


if __name__ == "__main__":
    main()
